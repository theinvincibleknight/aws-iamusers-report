import boto3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
from io import BytesIO
import logging

logger = logging.getLogger()
logger.setLevel(logging.INFO)

# S3 bucket configuration
S3_BUCKET = 'aws-artifact-collector'

# Account configurations with assume role ARNs
ACCOUNTS = {
    'dev': {
        'role_arn': 'arn:aws:iam::111111111111:role/inventory-readonly-role',
    },
    'uat': {
        'role_arn': 'arn:aws:iam::222222222222:role/inventory-readonly-role',
    },
    'prod': {
        'role_arn': 'arn:aws:iam::333333333333:role/inventory-readonly-role',
    },
    'oldprod': {
        'role_arn': 'arn:aws:iam::444444444444:role/inventory-readonly-role',
    },
    'network': {
        'role_arn': 'arn:aws:iam::555555555555:role/inventory-readonly-role',
    },
    'sharedservice': {
        'role_arn': 'arn:aws:iam::666666666666:role/inventory-readonly-role',
    },
}


def assume_role(role_arn, session_name='IAMInventorySession'):
    """Assume a role in a target account and return a boto3 session."""
    sts_client = boto3.client('sts')
    try:
        response = sts_client.assume_role(
            RoleArn=role_arn,
            RoleSessionName=session_name,
            DurationSeconds=3600
        )
        credentials = response['Credentials']
        session = boto3.Session(
            aws_access_key_id=credentials['AccessKeyId'],
            aws_secret_access_key=credentials['SecretAccessKey'],
            aws_session_token=credentials['SessionToken']
        )
        return session
    except Exception as e:
        logger.error(f"Error assuming role {role_arn}: {str(e)}")
        return None


def get_user_policies(iam_client, username):
    """Get all policies attached to a user - AWS managed, customer managed, and inline."""
    aws_managed_policies = []
    customer_managed_policies = []
    inline_policies = []

    try:
        # Get directly attached policies (both AWS and customer managed)
        paginator = iam_client.get_paginator('list_attached_user_policies')
        for page in paginator.paginate(UserName=username):
            for policy in page['AttachedPolicies']:
                policy_arn = policy['PolicyArn']
                if policy_arn.startswith('arn:aws:iam::aws:policy'):
                    aws_managed_policies.append(policy['PolicyName'])
                else:
                    customer_managed_policies.append(policy['PolicyName'])

        # Get inline policies directly on the user
        paginator = iam_client.get_paginator('list_user_policies')
        for page in paginator.paginate(UserName=username):
            inline_policies.extend(page['PolicyNames'])

        # Get policies inherited from groups
        paginator = iam_client.get_paginator('list_groups_for_user')
        for page in paginator.paginate(UserName=username):
            for group in page['Groups']:
                group_name = group['GroupName']

                # Group attached policies
                grp_paginator = iam_client.get_paginator('list_attached_group_policies')
                for grp_page in grp_paginator.paginate(GroupName=group_name):
                    for policy in grp_page['AttachedPolicies']:
                        policy_arn = policy['PolicyArn']
                        policy_display = f"{policy['PolicyName']} (via group: {group_name})"
                        if policy_arn.startswith('arn:aws:iam::aws:policy'):
                            aws_managed_policies.append(policy_display)
                        else:
                            customer_managed_policies.append(policy_display)

                # Group inline policies
                grp_inline_paginator = iam_client.get_paginator('list_group_policies')
                for grp_page in grp_inline_paginator.paginate(GroupName=group_name):
                    for pol_name in grp_page['PolicyNames']:
                        inline_policies.append(f"{pol_name} (via group: {group_name})")

    except Exception as e:
        logger.error(f"Error getting policies for user {username}: {str(e)}")

    return aws_managed_policies, customer_managed_policies, inline_policies


def get_access_key_details(iam_client, username):
    """Get access key details for a user including key age and last used info."""
    access_keys = []
    try:
        response = iam_client.list_access_keys(UserName=username)
        for key_metadata in response['AccessKeyMetadata']:
            key_id = key_metadata['AccessKeyId']
            status = key_metadata['Status']
            create_date = key_metadata['CreateDate']

            # Calculate key age (only meaningful for active keys)
            now = datetime.now(timezone.utc)
            key_age_days = (now - create_date).days

            # Get last used info
            last_used_info = 'Never'
            last_used_service = 'N/A'
            try:
                last_used_response = iam_client.get_access_key_last_used(AccessKeyId=key_id)
                access_key_last_used = last_used_response.get('AccessKeyLastUsed', {})
                if 'LastUsedDate' in access_key_last_used:
                    last_used_date = access_key_last_used['LastUsedDate']
                    last_used_info = last_used_date.strftime('%Y-%m-%d %H:%M:%S')
                    last_used_service = access_key_last_used.get('ServiceName', 'N/A')
            except Exception:
                pass

            access_keys.append({
                'key_id': key_id,
                'status': status,
                'create_date': create_date.strftime('%Y-%m-%d %H:%M:%S'),
                'key_age_days': key_age_days if status == 'Active' else 'N/A (Inactive)',
                'last_used': last_used_info,
                'last_used_service': last_used_service,
            })
    except Exception as e:
        logger.error(f"Error getting access keys for user {username}: {str(e)}")

    return access_keys


def get_last_activity(iam_client, username):
    """Get the last activity timestamp for a user (password last used)."""
    try:
        response = iam_client.get_user(UserName=username)
        if 'PasswordLastUsed' in response['User']:
            return response['User']['PasswordLastUsed'].strftime('%Y-%m-%d %H:%M:%S')
        return 'No console activity'
    except Exception as e:
        logger.error(f"Error getting last activity for {username}: {str(e)}")
        return 'Error'


def get_console_access_status(iam_client, username):
    """Check if a user has console access (login profile exists)."""
    try:
        iam_client.get_login_profile(UserName=username)
        return 'Enabled'
    except iam_client.exceptions.NoSuchEntityException:
        return 'Disabled'
    except Exception as e:
        logger.error(f"Error checking console access for {username}: {str(e)}")
        return 'Error'


def get_user_creation_time(iam_client, username):
    """Get user creation time."""
    try:
        response = iam_client.get_user(UserName=username)
        return response['User']['CreateDate'].strftime('%Y-%m-%d %H:%M:%S')
    except Exception as e:
        logger.error(f"Error getting creation time for {username}: {str(e)}")
        return 'Error'


def collect_iam_data(account_name, session):
    """Collect all IAM user data for a given account session."""
    iam_client = session.client('iam')
    users_data = []

    try:
        # List all IAM users
        paginator = iam_client.get_paginator('list_users')
        all_users = []
        for page in paginator.paginate():
            all_users.extend(page['Users'])

        logger.info(f"Found {len(all_users)} users in account: {account_name}")

        for user in all_users:
            username = user['UserName']
            logger.info(f"Processing user: {username} in account: {account_name}")

            # Get policies
            aws_managed, customer_managed, inline = get_user_policies(iam_client, username)

            # Get access keys
            access_keys = get_access_key_details(iam_client, username)

            # Get last activity
            last_activity = get_last_activity(iam_client, username)

            # Get console access
            console_access = get_console_access_status(iam_client, username)

            # Get creation time
            creation_time = get_user_creation_time(iam_client, username)

            # One row per access key; if user has no keys, still one row
            if access_keys:
                for key in access_keys:
                    users_data.append({
                        'Account': account_name,
                        'UserName': username,
                        'AWS Managed Policies': '\n'.join(aws_managed) if aws_managed else 'None',
                        'Customer Managed Policies': '\n'.join(customer_managed) if customer_managed else 'None',
                        'Inline Policies': '\n'.join(inline) if inline else 'None',
                        'Last Activity': last_activity,
                        'Access Key ID': key['key_id'],
                        'Access Key Status': key['status'],
                        'Active Key Age (Days)': key['key_age_days'],
                        'Access Key Last Used': key['last_used'],
                        'Access Key Last Used Service': key['last_used_service'],
                        'Access Key Creation Time': key['create_date'],
                        'Console Access': console_access,
                        'User Creation Time': creation_time,
                    })
            else:
                users_data.append({
                    'Account': account_name,
                    'UserName': username,
                    'AWS Managed Policies': '\n'.join(aws_managed) if aws_managed else 'None',
                    'Customer Managed Policies': '\n'.join(customer_managed) if customer_managed else 'None',
                    'Inline Policies': '\n'.join(inline) if inline else 'None',
                    'Last Activity': last_activity,
                    'Access Key ID': 'No Access Key',
                    'Access Key Status': 'N/A',
                    'Active Key Age (Days)': 'N/A',
                    'Access Key Last Used': 'N/A',
                    'Access Key Last Used Service': 'N/A',
                    'Access Key Creation Time': 'N/A',
                    'Console Access': console_access,
                    'User Creation Time': creation_time,
                })

    except Exception as e:
        logger.error(f"Error collecting IAM data for account {account_name}: {str(e)}")

    return users_data


def create_excel_report(all_data):
    """Create an Excel workbook from the collected IAM data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'IAM Users Report'

    # Define headers
    headers = [
        'Sr. No',
        'Account',
        'UserName',
        'AWS Managed Policies',
        'Customer Managed Policies',
        'Inline Policies',
        'Last Activity',
        'Access Key ID',
        'Access Key Status',
        'Active Key Age (Days)',
        'Access Key Last Used',
        'Access Key Last Used Service',
        'Access Key Creation Time',
        'Console Access',
        'User Creation Time',
    ]

    # Header styling
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    for row_idx, row_data in enumerate(all_data, 2):
        for col_idx, header in enumerate(headers, 1):
            if header == 'Sr. No':
                value = row_idx - 1  # Serial number starting from 1
            else:
                value = row_data.get(header, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border

            # Highlight inactive keys in red
            if header == 'Access Key Status' and row_data.get(header) == 'Inactive':
                cell.font = Font(color='FF0000')

            # Highlight active key age > 90 days in orange/bold
            if header == 'Active Key Age (Days)':
                val = row_data.get(header)
                if isinstance(val, int) and val > 90:
                    cell.font = Font(color='FF8C00', bold=True)

            # Highlight disabled console access in gray
            if header == 'Console Access' and row_data.get(header) == 'Disabled':
                cell.font = Font(color='808080')

    # Auto-adjust column widths (capped to keep rows compact with wrap_text)
    for col in range(1, len(headers) + 1):
        max_length = len(str(ws.cell(row=1, column=col).value))
        for row in range(2, min(ws.max_row + 1, 50)):  # Sample first 50 rows
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                lines = str(cell_value).split('\n')
                max_line_len = max(len(line) for line in lines)
                max_length = max(max_length, max_line_len)
        # Cap width at 35 to keep rows compact; wrap_text handles overflow
        adjusted_width = min(max_length + 2, 35)
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

    return wb


def upload_to_s3(workbook):
    """Save workbook to S3 with path: iam_users/<year>/<month>/iam_users_report_<year>_<month>_<HHMMSS>.xlsx"""
    now = datetime.now(timezone.utc)
    year = now.strftime('%Y')
    month = now.strftime('%m')
    time_str = now.strftime('%H%M%S')

    s3_key = f"iam_users/{year}/{month}/iam_users_report_{year}_{month}_{time_str}.xlsx"

    # Save workbook to bytes buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Upload to S3
    s3_client = boto3.client('s3')
    s3_client.put_object(
        Bucket=S3_BUCKET,
        Key=s3_key,
        Body=buffer.getvalue(),
        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    s3_path = f"s3://{S3_BUCKET}/{s3_key}"
    logger.info(f"Report uploaded to: {s3_path}")
    return s3_path


def lambda_handler(event, context):
    """Main Lambda handler function."""
    logger.info("Starting IAM Users Inventory Report generation...")

    all_data = []

    for account_name, account_config in ACCOUNTS.items():
        role_arn = account_config['role_arn']
        logger.info(f"Processing account: {account_name} ({role_arn})")

        session = assume_role(role_arn)
        if session is None:
            logger.error(f"Skipping account {account_name} - could not assume role")
            continue

        account_data = collect_iam_data(account_name, session)
        all_data.extend(account_data)
        logger.info(f"Collected {len(account_data)} records from account: {account_name}")

    if not all_data:
        logger.warning("No data collected from any account!")
        return {
            'statusCode': 200,
            'body': 'No IAM user data collected from any account.'
        }

    logger.info(f"Total records collected: {len(all_data)}")

    # Create Excel report
    workbook = create_excel_report(all_data)

    # Upload to S3
    s3_path = upload_to_s3(workbook)

    return {
        'statusCode': 200,
        'body': f'IAM Users Inventory Report generated successfully. File: {s3_path}'
    }
